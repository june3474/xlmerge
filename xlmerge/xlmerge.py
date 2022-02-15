# -*- coding: utf-8 -*-

import sys
import os
import openpyxl as pyxl
import xlwings as xw
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtWidgets import QApplication, QDialog, QTableWidgetItem, QFileDialog, QMessageBox
import win32gui
from win32con import SW_MAXIMIZE
from adodbapi.adodbapi import pywintypes
from ui import Ui_Dialog
from utils import CellAddr, MyConverter, list_strip, xl_col_to_name
from addin import Addin


class HeaderSelector(QDialog):
    def __init__(self, parent=None, rows=10, cols=10):
        """Class for Selecting the header(title) row.

        ``HeaderSelector`` opens Excel first; let the user select Excel files to merge;
        read and show a part of the file so that the user selects the header row.
        Selected headers are put the in the sheet named "Headers".
        When finished collecting headers, HeaderSelector put a button at the end that
        triggers merging.

        Args:
            parent (:obj:`QWidget`, optional): Qt widget that HeaderSelector instance
                belongs to. if parent is None, HeaderSelector becomes the top container.
            rows (int, optional): The number of rows(default 10) to show in QTableWidget.
            cols (int, optional): The number of columns(default 10) to show in QTableWidget.

        """

        super().__init__(parent)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)
        # self.setAttribute(Qt.WA_ShowWithoutActivating, True)
        self.rows = 1000 if cols > 1000 else rows  # limit number of rows to show
        self.cols = 26 if cols > 26 else cols  # 'A' ~ 'Z'
        self._init_table()

        self.files = None
        self.sheets = []
        self.current_sheet = 0
        self.xlmerge_path = self._find_xlmerge()
        self.header_sheet = self.open_excel()
        self.excel_pid = self.header_sheet.book.app.pid

        MyConverter.register('myconv')

    def _init_table(self):
        """Initialize table(:obj:`QTableWidget`)

        This methode set the dimension of the table, and set the column heading.
        Some properties like borders, and font are defined in the stylesheet
        in ``ui.py``.

        """

        self.ui.table.setRowCount(self.rows)
        self.ui.table.setColumnCount(self.cols)
        self.ui.table.setSortingEnabled(False)
        # chr(65) == 'A'
        col_label = [chr(65+i) for i in range(self.cols)]
        self.ui.table.setHorizontalHeaderLabels(col_label)

    def _find_xlmerge(self):
        """Find the location of ``xlmerge.exe``

        Returns (str): Full path of ``xlmerge.exe``

        """
        if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):  # bundled exe
            return os.path.join(sys._MEIPASS, 'xlmerge.exe')
        else:  # run .py directly or package exe
            return os.path.join(os.path.dirname(sys.executable), 'Scripts', 'xlmerge.exe')

    def open_excel(self):
        """Open a new Excel instance and prepare a workbook and a worksheet.

        In addition, this method inserts a macro module which contains macros defined in
        ``addin.template``. This is necessary because the newly created workbook by this
        method does not load ``xlmerge.xlam`` in ``%UserProfile%\\AppData\\Roaming\\Microsoft\\AddIns``

        """

        excel = xw.App(visible=True, add_book=False)
        # Add a workbook with only one sheet.
        # Use -4167 directly instead of enum to avoid adding a reference to
        # 'Microsoft Visual Basic for Applications Extensibility 5.3'.
        excel.api.Application.Workbooks.Add(-4167)
        win32gui.ShowWindow(excel.hwnd, SW_MAXIMIZE)
        win32gui.SetForegroundWindow(excel.hwnd)
        wb = excel.books.active

        # Insert macro
        Addin().insert_macro(wb, self.xlmerge_path)

        ws = wb.sheets.active
        ws.name = 'Headers'
        return ws

    def select_files(self):
        """Open :class:`QFileDialog` to select files to merge"""

        # getOpenFilenames returns a tuple: ([list_of_filepath], filter)
        files, _ = QFileDialog.getOpenFileNames(None, 'xlmerge - 병합할 파일 선택', '',
                                                "Excel files (*.xls *.xlsx *.xlsm);;All files (*)")
        self.files = files
        return files

    def read_files(self):  # openpyxl
        """Read rows as many as ``self.rows`` from selected files.

        Then, HeaderSelector fills the list variable ``self.sheets`` with the data read from files.

        """

        if not self.files:
            return False
        # Change cursor to hour glass
        self.header_sheet.api.Application.Cursor = 2
        for f in self.files:
            try:
                wb = pyxl.load_workbook(filename=f, read_only=True)
            except:
                ret = QMessageBox.warning(self, "파일 읽기 에러",
                                          "{}을 읽을 수 없습니다.\n다음 파일로 진행하시겠습니까?".format(f),
                                          QMessageBox.Yes | QMessageBox.No,
                                          QMessageBox.Yes)
                if ret == 0x00004000:  # QMessageBox::Yes
                    continue
                else:
                    # Restore default cursor
                    self.header_sheet.api.Application.Cursor = -4143
                    return False

            for name in wb.sheetnames:
                ws = wb[name]
                data = []
                for r in ws.iter_rows(max_row=self.rows, values_only=True):
                    data.append(r)
                # {'file': 'filename', 'sheet':'sheetname',
                # 'data': [(val_A1, val_B1...), (val_B1, val_B2...) ... ]}
                self.sheets.append(dict(file=f, sheet=name, data=data))

            wb.close()

        self.header_sheet.api.Application.Cursor = -4143
        return True

    def load_sheet(self):
        """Load sheet info into HeaderSelector UI.

        Precondition: ``self.sheets`` is NOT empty. i.e., len(self.sheets) >= 1.

        """

        # if this is the last sheets, disable the next button.
        if self.current_sheet == len(self.sheets) - 1:
            self.ui.btnNext.setDisabled(True)

        # [{'file':'filename', 'sheet':'sheetname', 'data':[(val_A1, val_B1...), ... ]}]
        sheet = self.sheets[self.current_sheet]
        # Update label
        file_text = '{filename}  ({current}/{total})'.format(filename=os.path.split(sheet.get('file'))[1],
                                                             current=self.files.index(sheet.get('file'))+1,
                                                             total=len(self.files))
        self.ui.labelFile.setText(file_text)
        sheet_text = '{sheetname}  ({current}/{total})'.format(sheetname=sheet.get('sheet'),
                                                               current=self.current_sheet+1,
                                                               total=len(self.sheets))
        self.ui.labelSheet.setText(sheet_text)

        # Update progressbar
        self.ui.progressBar.setValue(round(self.current_sheet / len(self.sheets) * 100))

        # Update table
        # Clear table if necessary
        if self.current_sheet:  # if not first sheet, i.e., not 0
            self.ui.table.clearContents()
        # Fill table
        # QTableWidget uses 0-based index
        for row, datum in enumerate(sheet.get('data')):
            for col, d in enumerate(datum):
                if col == self.cols:
                    break
                d = str(d) if type(d) is int else d
                item = QTableWidgetItem(d)
                self.ui.table.setItem(row, col, item)

    @staticmethod
    def _last_row(ws):  # xlwings
        """Returns the first empty row of the given worksheet.

        Args:
            ws (:obj:`xlwings.Sheet`): `xlwings.Sheet` representing Excel worksheet

        Returns (int):
            index of the first empty row

        """

        # The exact last row can be calculated by 'ws.cells.last_cell.row'.
        # The maximum row count of HeadSelector is 1000.
        last = ws.range('A1000').end('up').row
        if ws.range('A1').value:
            last += 1
        return last

    def write_to_excel(self):  # xlwings
        """Write the selected row(header) into the Excel sheet

        In addition to the data contained in the worksheet, This methode add following info:
            - fullpath of the workbook
            - sheet name
            - header row index

        """

        sheet = self.sheets[self.current_sheet]
        selected_index = self.ui.table.selectedIndexes()
        selected_row = selected_index[0].row() + 1 if selected_index else None
        # In case that user did not select a header row
        if not selected_row:
            return

        sheet['row'] = selected_row
        header = list_strip(sheet['data'][selected_row-1])
        # Add filepath, sheetname, index of the header row
        value = [sheet['file'], sheet['sheet'], sheet['row'], *header]
        last_row = str(self._last_row(self.header_sheet))
        addr = 'A' + last_row

        # Write to excel
        self.header_sheet.range(addr).options('myconv', empty='').value = value
        # Paint yellow the first 3 columns for filepath, sheetname, header row index
        self.header_sheet.range(addr, addr.replace('A', 'C')).color = '#ffff00'
        # Put validation
        macro = self.header_sheet.book.macro('put_validation')
        where = 'D' + last_row + ':' + xl_col_to_name(3+len(header)-1) + last_row
        macro(where, ','.join(header))

        self.header_sheet.autofit('c')  # columns only

    def kill_excel(self):
        """Close Excel instance"""

        if self.excel_pid:
            try:
                xw.apps[self.excel_pid].quit()  # quit without saving
            except KeyError:  # already closed by user
                pass
        self.excel_pid = None

    @pyqtSlot()
    def on_btnNext_clicked(self):
        """다음 button handler"""

        self.write_to_excel()
        self.current_sheet += 1
        self.load_sheet()

    @pyqtSlot()
    def on_btnFinish_clicked(self):
        """종료 button handler"""

        self.write_to_excel()
        # if 'Header' sheet is empty
        if self.header_sheet.used_range.count == 1 and self.header_sheet.used_range.value is None:
            self.kill_excel()
        else:
            # put a button at the end
            macro = self.header_sheet.book.macro('insert_button')
            insert_position = '{}{}:{}{}'.format('A', self._last_row(self.header_sheet) + 1,
                                                 'A', self._last_row(self.header_sheet) + 2)
            macro('병합 시작', insert_position)

        self.close()
        win32gui.SetForegroundWindow(self.header_sheet.book.app.hwnd)

    def run(self):
        """method to facilitate easy execution"""

        if not self.select_files():
            self.kill_excel()
            return False
        if not self.read_files():
            return False
        if not self.sheets:
            return False
        self.load_sheet()
        self.show()
        return True


class Merger:
    def __init__(self):
        """Class for merging sheets

        This class performs merging based on the sheet HeaderSelector creates.

        """

        #  merge gets run only through the button click on the 'Headers' sheet
        self.excel = xw.apps.active
        self.wb = self.excel.books.active
        self.header_sheet = self.wb.sheets['Headers']
        self.merge_sheet = self.create_merge_sheet()
        self.write_point = CellAddr()
        # Register MyConverter alias
        MyConverter.register('myconv')
        self.show_wait_message()
        self.headers = self.read_headers()  # list of each row in header_sheet

    def show_wait_message(self):
        """Show wait message"""

        msg = "병합 중... 시간이 좀(?) 걸릴 수 있습니다."
        self.merge_sheet.range('A1').value = msg

    def read_headers(self):  # xlwingsw
        """Read cells on the ``Headers`` sheet.

        Returns (list):
            list of cells read

        """

        headers = []
        for r in self.header_sheet.range('A1').current_region.rows:
            headers.append([c.options('myconv', empty='').value for c in r])

        return headers

    def create_merge_sheet(self):  # xlwings
        """Find or create a new sheet named ``Merge`` for merging"""

        try:
            return self.wb.sheets['Merge']  # if already exists
        except pywintypes.com_error:
            return self.wb.sheets.add(name='Merge', after=self.header_sheet)

    def read_cols(self, sheet, row, col):  # openpyxl
        """Read data by column"""

        # openpyxl does not provide iter_cols() in read-only mode
        return [r[0].value for r in sheet.iter_rows(min_row=row, min_col=col, max_col=col)]

    def write_cols(self, data, celladdr=None):  # xlwings
        """Write data into ``Merge`` sheet by column"""

        if not celladdr:
            celladdr = self.write_point.addr
        self.merge_sheet.range(celladdr).options('myconv', empty='', transpose=True).value = data

    def run(self):  # openpyxl & xlwings
        """method to facilitate easy execution"""

        self.header_sheet.api.Application.Cursor = 2
        self.excel.screen_updating = False
        # pre process, pop 3 cols
        for header in self.headers:
            filepath = header.pop(0)
            sheetname = header.pop(0)
            header_row = header.pop(0)
            # read file
            wb = pyxl.load_workbook(filepath, read_only=True)
            ws = wb[sheetname]
            source_header = [c.value for r in ws.iter_rows(min_row=header_row, max_row=header_row) for c in r]

            # Start writing from column 'B', save column 'A' for filename
            # Todo: tidy up this ugly column manipulation
            self.write_point.increase_col()
            for item in header:
                # if item is empty, do nothing and proceed to the next item
                if item:
                    # locate col's index in the source excel file
                    col = source_header.index(item) + 1
                    data = self.read_cols(ws, header_row, col)
                    self.write_cols(data)
                self.write_point.increase_col()

            # Insert filename and sheetname in the first column
            _, filename = os.path.split(filepath)
            fileNsheet = filename + '-' + sheetname
            data = [fileNsheet] * len(data)  # closure
            self.write_cols(data, (self.write_point.row, 1))
            wb.close()

            header_row = str(self.write_point.row)
            self.merge_sheet.range(header_row + ':' + header_row).color = '#ffff00'
            self.write_point.set(self.write_point.row + len(data), 1)

        self.merge_sheet.autofit('c')
        self.excel.screen_updating = True
        self.merge_sheet.activate()
        self.header_sheet.api.Application.Cursor = -4143

    def remove_macros(self):
        """Remove macro module from the workbook

        By removing macro module, our workbook can be saved as normal xlsx file

        """

        self.merge_sheet.api.VBProject.VBComponents.Remove("xlmerge")


def main():
    # sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'lib'))
    # sys.path.append(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'packages'))

    if len(sys.argv) > 1 and sys.argv[1] == 'merge':
        merger = Merger()
        merger.run()
        sys.exit()
        # merger.remove_macros()

    app = QApplication(sys.argv)  # Not merge,, i.e., header select
    headerSelector = HeaderSelector()
    if not headerSelector.run():  # Either no file selected or error in file reading
        headerSelector.close()
        sys.exit()

    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
